"""
Genera el archivo .xlsx plantilla que los editores rellenan y suben.
Ejecutar directamente para regenerar: python -m app.services.excel_template
"""
from __future__ import annotations
from pathlib import Path
import openpyxl
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

MAG   = "C8006B"
NAVY  = "1A1A2E"
LIGHT = "F5F5F5"
WHITE = "FFFFFF"
GRAY  = "E0E0E0"

# ── Style helpers ─────────────────────────────────────────────────────────────

def _h(ws, row, col, value, bold=False, size=11, bg=None, fg="000000", wrap=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=bold, size=size, color=fg,
                     name="DM Sans" if bold else "Inter")
    cell.alignment = Alignment(wrap_text=wrap, vertical="center")
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    return cell


def _header_row(ws, row, labels, bg=NAVY, fg=WHITE):
    for col, label in enumerate(labels, 1):
        _h(ws, row, col, label, bold=True, size=9, bg=bg, fg=fg)
    ws.row_dimensions[row].height = 22


def _thin_border():
    side = Side(style="thin", color=GRAY)
    return Border(left=side, right=side, top=side, bottom=side)


def _col_widths(ws, widths: dict):
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


# ── Sheets ────────────────────────────────────────────────────────────────────

def _sheet_metadata(wb):
    ws = wb.create_sheet("METADATA")
    ws.sheet_view.showGridLines = False

    # Title
    _h(ws, 1, 1, "GUÍA DISCOOLVER — METADATA", bold=True, size=14, fg=MAG)
    _h(ws, 2, 1, "Rellena los campos en la columna B. No borres las etiquetas.", size=9, fg="666666")

    fields = [
        ("city",              "Ciudad",               "BARCELONA"),
        ("year",              "Año (2 dígitos)",       "26"),
        ("edition",           "Nombre edición",        "Edición Barcelona 2026"),
        ("director",          "Director / Editor",     "Carlos Jacoste"),
        ("director_role",     "Cargo director",        "CEO & Fundador — discoolver"),
        ("collection",        "Colección",             "estandar"),
        ("accent_color",      "Color accent (HEX)",    "#C8006B"),
        ("status",            "Estado",                "draft"),
        ("cover_headline1",   "Portada — Titular 1",   "INSPIRING"),
        ("cover_headline2",   "Portada — Titular 2",   "the World"),
        ("cover_tagline",     "Portada — Tagline",     "coolest places in the world"),
        ("headline_align",    "Alineación titular",    "right"),
        ("directors_letter",  "Carta del director",    ""),
        ("director_pull_quote","Pull-quote carta",     ""),
        ("mission_text",      "Texto misión",          ""),
        ("persona_name",      "Persona del Año — Nombre", ""),
        ("persona_tagline",   "Persona del Año — Tagline",""),
        ("persona_origen",    "Persona del Año — Origen", ""),
        ("persona_disciplina","Persona del Año — Disciplina",""),
        ("persona_bio",       "Persona del Año — Bio", ""),
        ("persona_quote",     "Persona del Año — Cita", ""),
    ]

    _header_row(ws, 4, ["Campo (no editar)", "Valor", "Descripción"])

    for i, (key, label, example) in enumerate(fields, 5):
        _h(ws, i, 1, key, bold=False, size=9, bg=LIGHT, fg="444444")
        cell = ws.cell(row=i, column=2, value=example)
        cell.font = Font(size=10)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        _h(ws, i, 3, label, size=9, fg="888888")
        ws.row_dimensions[i].height = 20

    # Collection dropdown
    dv = DataValidation(
        type="list",
        formula1='"estandar,nomadas-digitales,ocio-nocturno,gastronomia,influencers,luxury,custom"',
        showDropDown=False,
    )
    ws.add_data_validation(dv)
    dv.add(ws["B10"])

    # Status dropdown
    dv2 = DataValidation(type="list", formula1='"draft,review,published"', showDropDown=False)
    ws.add_data_validation(dv2)
    dv2.add(ws["B14"])

    _col_widths(ws, {"A": 26, "B": 42, "C": 30})
    return ws


def _sheet_recomendados(wb):
    ws = wb.create_sheet("RECOMENDADOS")
    ws.sheet_view.showGridLines = False

    _h(ws, 1, 1, "RECOMENDADOS — Fichas de lugares", bold=True, size=13, fg=MAG)
    _h(ws, 2, 1,
       "Cada fila = 1 ficha. La columna SECTION define en qué página aparece. "
       "Deja PHOTO_URL vacío si subirás la foto desde el editor.",
       size=9, fg="666666", wrap=True)
    ws.row_dimensions[2].height = 28

    headers = [
        "SECTION", "SUBCATEGORY", "BADGE", "NAME", "TAGLINE",
        "DESCRIPTION", "WEB", "ADDRESS", "DISCOOLVER_URL",
        "PHOTO_URL", "SORT_ORDER", "ENABLED",
    ]
    _header_row(ws, 4, headers)

    # Section dropdown
    sections = (
        "restaurantes,gastronomia_bcn,fiesta,ocio_eventos,"
        "arte_exposiciones,experiencias,alojamientos,shopping,persona_recom"
    )
    dv = DataValidation(type="list", formula1=f'"{sections}"', showDropDown=False)
    ws.add_data_validation(dv)
    for row in range(5, 205):
        dv.add(ws.cell(row=row, column=1))

    # Sample rows
    samples = [
        ("restaurantes",       "Exclusivo",    "LOCAL",   "71 Oyster Bar",       "Bar de ostras y cócteles", "El mejor bar de ostras de Barcelona.", "https://example.com", "C/ Enric Granados 71", "https://discoolver.com/bcn/71-oyster-bar", "", 1, "TRUE"),
        ("restaurantes",       "Trendy",       "",        "Disfrutar",           "Alta cocina de vanguardia",  "Top 3 en The World's 50 Best.", "https://disfrutarbarcelona.com", "C/ de Villarroel 163", "", "", 2, "TRUE"),
        ("fiesta",             "BARES DE COPAS","",       "Blackwell Rum Bar",   "Jamaica. Una cueva volcánica.", "Bar en el acantilado.", "https://thecaveshotel.com", "West End, Negril, Jamaica", "", "", 1, "TRUE"),
        ("ocio_eventos",       "",             "",        "Primavera Sound",     "Festival destacado del año", "El festival indie más importante de Europa.", "", "Parc del Fòrum, Barcelona", "", "", 1, "TRUE"),
        ("influencers",        "",             "",        "",                    "", "", "", "", "", "", 1, "TRUE"),
    ]
    for i, row_data in enumerate(samples, 5):
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.font = Font(size=9)
            cell.alignment = Alignment(wrap_text=False, vertical="center")
        ws.row_dimensions[i].height = 18

    _col_widths(ws, {
        "A": 18, "B": 16, "C": 10, "D": 26,
        "E": 24, "F": 40, "G": 28, "H": 30,
        "I": 30, "J": 32, "K": 10, "L": 8,
    })
    return ws


def _sheet_influencers(wb):
    ws = wb.create_sheet("INFLUENCERS")
    ws.sheet_view.showGridLines = False

    _h(ws, 1, 1, "LOCAL INFLUENCERS", bold=True, size=13, fg=MAG)

    headers = [
        "NAME", "HANDLE", "PLATFORM", "CITY",
        "DESCRIPTION", "FOLLOWERS", "ENGAGEMENT",
        "CATEGORY_1", "CATEGORY_2", "CATEGORY_3",
        "PHOTO_URL", "ENABLED",
    ]
    _header_row(ws, 3, headers)

    platform_dv = DataValidation(
        type="list",
        formula1='"Instagram,TikTok,YouTube,Twitter,Twitch"',
        showDropDown=False,
    )
    ws.add_data_validation(platform_dv)
    for row in range(4, 54):
        platform_dv.add(ws.cell(row=row, column=3))

    samples = [
        ("Alba Díaz",    "@albadiazwilstermann", "Instagram", "Madrid, España",    "Creadora de contenido de moda, lifestyle y viajes.", "1.2M",  "8.4%",  "MODA",    "LIFESTYLE", "VIAJES",  "", "TRUE"),
        ("Ibai Llanos",  "@ibaillanos",           "TikTok",   "Barcelona, España", "Streamer y creador de contenido referente.", "9.8M",  "12.1%", "GAMING",  "HUMOR",     "DEPORTES","", "TRUE"),
        ("Dulceida",     "@dulceida",             "YouTube",  "Barcelona, España", "Pionera del blogging y las redes sociales en España.", "3.1M",  "5.2%",  "MODA",    "BEAUTY",    "VIAJES",  "", "TRUE"),
    ]
    for i, row_data in enumerate(samples, 4):
        for col, val in enumerate(row_data, 1):
            ws.cell(row=i, column=col, value=val).font = Font(size=9)
        ws.row_dimensions[i].height = 18

    _col_widths(ws, {
        "A": 22, "B": 24, "C": 12, "D": 20,
        "E": 38, "F": 10, "G": 10,
        "H": 14, "I": 14, "J": 14,
        "K": 32, "L": 8,
    })
    return ws


def _sheet_persona(wb):
    ws = wb.create_sheet("PERSONA_DEL_ANO")
    ws.sheet_view.showGridLines = False

    _h(ws, 1, 1, "PERSONA DEL AÑO — Timeline y Recomendados", bold=True, size=13, fg=MAG)

    # Timeline
    _h(ws, 3, 1, "TIMELINE", bold=True, size=10, fg=NAVY)
    _header_row(ws, 4, ["YEAR", "ITEM_1", "ITEM_2", "ITEM_3", "ITEM_4"], bg=NAVY)

    tl_samples = [
        ("2016", "EP debut independiente", '"Antes de morirme" ft. C. Tangana', "", ""),
        ("2018", '"El Mal Querer" — álbum del año', "Grammy Latino mejor álbum alternativo", "Latin Grammy álbum del año", ""),
        ("2022", 'Álbum "MOTOMAMI" — disco del año', "Platino en 14 países", "TIME 100 Most Influential People", ""),
        ("2026", "Nuevo álbum anunciado", "★ Persona del Año discoolver BCN", "", ""),
    ]
    for i, row_data in enumerate(tl_samples, 5):
        for col, val in enumerate(row_data, 1):
            ws.cell(row=i, column=col, value=val).font = Font(size=9)

    # Awards
    _h(ws, 12, 1, "RECONOCIMIENTOS (uno por celda)", bold=True, size=10, fg=NAVY)
    _header_row(ws, 13, ["AWARD_1", "AWARD_2", "AWARD_3", "AWARD_4", "AWARD_5", "AWARD_6"], bg=NAVY)
    for col, val in enumerate(["Grammy Latino ×4", "MTV EMA", "NME Award", "Brit Award", "Premio Nac. Músicas Actuales", "TIME 100"], 1):
        ws.cell(row=14, column=col, value=val).font = Font(size=9)

    # Recomendados persona
    _h(ws, 17, 1, "RECOMENDADOS DE LA PERSONA DEL AÑO (en su ciudad)", bold=True, size=10, fg=NAVY)
    headers_rec = ["NAME", "CATEGORY (badge)", "DESCRIPTION", "ADDRESS", "DISCOOLVER_URL", "PHOTO_URL"]
    _header_row(ws, 18, headers_rec, bg=NAVY)

    _col_widths(ws, {"A": 20, "B": 22, "C": 40, "D": 30, "E": 30, "F": 30})
    return ws


def _sheet_awards(wb):
    ws = wb.create_sheet("AYUDA")
    ws.sheet_view.showGridLines = False

    _h(ws, 1, 1, "GUÍA DE USO — Valores válidos por columna", bold=True, size=13, fg=MAG)

    help_data = [
        ("SECTION (hoja RECOMENDADOS)", "restaurantes | gastronomia_bcn | fiesta | ocio_eventos | arte_exposiciones | experiencias | alojamientos | shopping | persona_recom"),
        ("SUBCATEGORY (restaurantes)", "Exclusivo | Trendy | Tradicional | WOW | Pet Friendly | Food Truck"),
        ("SUBCATEGORY (fiesta)",       "BARES DE COPAS | SPEAKEASY | CLUBBING"),
        ("SUBCATEGORY (arte)",         "EXPOSICIONES | TEATRO Y ESPECTÁCULOS"),
        ("SUBCATEGORY (experiencias)", "EXPERIENCIAS WOW | ACTIVIDADES COOL"),
        ("SUBCATEGORY (alojamientos)", "ALOJAMIENTOS TOP | ESPACIOS ÚNICOS"),
        ("SUBCATEGORY (shopping)",     "COOL PLACES & PRODUCTS | PRODUCTOS DESTACADOS"),
        ("BADGE (alojamientos)",       "LUXURY | BOUTIQUE | APART | RESORT"),
        ("BADGE (shopping)",           "LOCAL | ICÓNICO | COOL | EXCLUSIVO"),
        ("BADGE (experiencias)",       "WOW"),
        ("PLATFORM (influencers)",     "Instagram | TikTok | YouTube | Twitter | Twitch"),
        ("COLLECTION (metadata)",      "estandar | nomadas-digitales | ocio-nocturno | gastronomia | influencers | luxury | custom"),
        ("STATUS (metadata)",          "draft | review | published | archived"),
        ("ENABLED",                    "TRUE | FALSE"),
    ]

    _header_row(ws, 3, ["Columna", "Valores permitidos"])
    for i, (col, values) in enumerate(help_data, 4):
        _h(ws, i, 1, col, bold=True, size=9, bg=LIGHT)
        cell = ws.cell(row=i, column=2, value=values)
        cell.font = Font(size=9)
        cell.alignment = Alignment(wrap_text=True)
        ws.row_dimensions[i].height = 24

    _col_widths(ws, {"A": 30, "B": 80})
    return ws


# ── Build & save ──────────────────────────────────────────────────────────────

def build_template(output_path: Path | None = None) -> Path:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    _sheet_metadata(wb)
    _sheet_recomendados(wb)
    _sheet_influencers(wb)
    _sheet_persona(wb)
    _sheet_awards(wb)

    if output_path is None:
        output_path = Path(__file__).resolve().parent.parent.parent / "static" / "discoolver-guide-template.xlsx"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


if __name__ == "__main__":
    path = build_template()
    print(f"Template saved → {path}")
