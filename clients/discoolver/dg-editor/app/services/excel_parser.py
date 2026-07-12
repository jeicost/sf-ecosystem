"""
Parses the Discoolver guide Excel template → list of Pydantic models
ready for bulk insert into PostgreSQL.
"""
from __future__ import annotations
import re
from io import BytesIO
from typing import Any

import openpyxl

from app.models.guide_v2 import (
    GuideCreate, GuideUpdate, ItemCreate, ItemType,
    GuideCollection, StatKPI, CriteriaItem,
)


# ── Helpers ───────────────────────────────────────────────────────────────────

def _cell(ws, row: int, col: int) -> Any:
    val = ws.cell(row=row, column=col).value
    if val is None:
        return None
    if isinstance(val, str):
        val = val.strip()
        return val if val else None
    return val


def _bool(val) -> bool:
    if val is None:
        return True
    return str(val).upper() in ("TRUE", "1", "SI", "SÍ", "YES")


def _row_as_dict(ws, row: int, headers: list[str]) -> dict:
    return {
        h: _cell(ws, row, col)
        for col, h in enumerate(headers, 1)
    }


# ── Sheet parsers ─────────────────────────────────────────────────────────────

def _parse_metadata(wb) -> tuple[GuideCreate, GuideUpdate]:
    ws = wb["METADATA"]
    data: dict[str, Any] = {}
    for row in ws.iter_rows(min_row=5, values_only=True):
        if row[0] and row[1] is not None:
            data[str(row[0]).strip()] = str(row[1]).strip() if row[1] != "" else None

    def get(key, default=None):
        return data.get(key) or default

    create = GuideCreate(
        city=get("city", "CIUDAD"),
        year=get("year", "26"),
        edition=get("edition"),
        director=get("director", "Carlos Jacoste"),
        director_role=get("director_role", "CEO & Fundador — discoolver"),
        collection=GuideCollection(get("collection", "estandar")),
        accent_color=get("accent_color"),
    )

    update = GuideUpdate(
        cover_headline1=get("cover_headline1"),
        cover_headline2=get("cover_headline2"),
        cover_tagline=get("cover_tagline"),
        headline_align=get("headline_align"),
        directors_letter=get("directors_letter"),
        director_pull_quote=get("director_pull_quote"),
        mission_text=get("mission_text"),
        persona_name=get("persona_name"),
        persona_tagline=get("persona_tagline"),
        persona_origen=get("persona_origen"),
        persona_disciplina=get("persona_disciplina"),
        persona_bio=get("persona_bio"),
        persona_quote=get("persona_quote"),
    )
    return create, update


def _parse_recomendados(wb) -> list[ItemCreate]:
    ws = wb["RECOMENDADOS"]
    items = []
    headers = [
        "SECTION", "SUBCATEGORY", "BADGE", "NAME", "TAGLINE",
        "DESCRIPTION", "WEB", "ADDRESS", "DISCOOLVER_URL",
        "PHOTO_URL", "SORT_ORDER", "ENABLED",
    ]

    for row in range(5, ws.max_row + 1):
        d = _row_as_dict(ws, row, headers)
        if not d.get("NAME") and not d.get("SECTION"):
            continue
        if not d.get("NAME"):
            continue

        section = (d.get("SECTION") or "restaurantes").strip()

        # Events (ocio_eventos) get item_type = event
        if section == "ocio_eventos":
            item_type = ItemType.event
        elif section == "persona_recom":
            item_type = ItemType.persona_recom
        else:
            item_type = ItemType.recomendado

        try:
            sort_order = int(d.get("SORT_ORDER") or 0)
        except (ValueError, TypeError):
            sort_order = 0

        items.append(ItemCreate(
            item_type=item_type,
            section=section,
            subcategory=d.get("SUBCATEGORY"),
            badge=d.get("BADGE"),
            name=str(d["NAME"]),
            tagline=d.get("TAGLINE"),
            description=d.get("DESCRIPTION"),
            web=d.get("WEB"),
            address=d.get("ADDRESS"),
            discoolver_url=d.get("DISCOOLVER_URL"),
            photo_url=d.get("PHOTO_URL"),
            sort_order=sort_order,
            enabled=_bool(d.get("ENABLED")),
        ))
    return items


def _parse_influencers(wb) -> list[ItemCreate]:
    ws = wb["INFLUENCERS"]
    items = []
    headers = [
        "NAME", "HANDLE", "PLATFORM", "CITY",
        "DESCRIPTION", "FOLLOWERS", "ENGAGEMENT",
        "CATEGORY_1", "CATEGORY_2", "CATEGORY_3",
        "PHOTO_URL", "ENABLED",
    ]

    for row in range(4, ws.max_row + 1):
        d = _row_as_dict(ws, row, headers)
        if not d.get("NAME"):
            continue

        stats = []
        if d.get("FOLLOWERS"):
            stats.append(StatKPI(num=str(d["FOLLOWERS"]), label="Seguidores"))
        if d.get("ENGAGEMENT"):
            stats.append(StatKPI(num=str(d["ENGAGEMENT"]), label="Engagement"))

        categories = [
            c for c in [d.get("CATEGORY_1"), d.get("CATEGORY_2"), d.get("CATEGORY_3")]
            if c
        ]

        items.append(ItemCreate(
            item_type=ItemType.influencer,
            section="influencers",
            name=str(d["NAME"]),
            description=d.get("DESCRIPTION"),
            photo_url=d.get("PHOTO_URL"),
            handle=d.get("HANDLE"),
            platform=d.get("PLATFORM"),
            city=d.get("CITY"),
            stats=stats,
            categories=categories,
            enabled=_bool(d.get("ENABLED")),
        ))
    return items


def _parse_persona(wb) -> tuple[list[ItemCreate], dict]:
    """Returns (timeline_items + recomendados, persona_extra_update)."""
    ws = wb["PERSONA_DEL_ANO"]
    items = []
    update_extra: dict[str, Any] = {}

    # Timeline rows 5-10 (approximate)
    for row in range(5, 12):
        year_val = _cell(ws, row, 1)
        if not year_val:
            break
        tl_items = [
            str(_cell(ws, row, c))
            for c in range(2, 6)
            if _cell(ws, row, c)
        ]
        items.append(ItemCreate(
            item_type=ItemType.timeline,
            section="persona_timeline",
            name=str(year_val),
            timeline_year=str(year_val),
            timeline_items=tl_items,
        ))

    # Awards row 14
    awards = [
        str(_cell(ws, 14, c))
        for c in range(1, 7)
        if _cell(ws, 14, c)
    ]
    if awards:
        update_extra["persona_awards"] = awards

    # Persona recomendados rows 19+
    rec_headers = ["NAME", "BADGE", "DESCRIPTION", "ADDRESS", "DISCOOLVER_URL", "PHOTO_URL"]
    for row in range(19, ws.max_row + 1):
        d = _row_as_dict(ws, row, rec_headers)
        if not d.get("NAME"):
            continue
        items.append(ItemCreate(
            item_type=ItemType.persona_recom,
            section="persona_recom",
            name=str(d["NAME"]),
            badge=d.get("BADGE"),
            description=d.get("DESCRIPTION"),
            address=d.get("ADDRESS"),
            discoolver_url=d.get("DISCOOLVER_URL"),
            photo_url=d.get("PHOTO_URL"),
        ))

    return items, update_extra


# ── Main entry point ──────────────────────────────────────────────────────────

class ExcelParseResult:
    def __init__(self):
        self.guide_create: GuideCreate | None = None
        self.guide_update: GuideUpdate | None = None
        self.items: list[ItemCreate] = []
        self.persona_extra: dict = {}
        self.errors: list[str] = []
        self.warnings: list[str] = []


def parse_excel(file_bytes: bytes) -> ExcelParseResult:
    result = ExcelParseResult()

    try:
        wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    except Exception as e:
        result.errors.append(f"No se pudo abrir el archivo Excel: {e}")
        return result

    required_sheets = {"METADATA", "RECOMENDADOS", "INFLUENCERS", "PERSONA_DEL_ANO"}
    missing = required_sheets - set(wb.sheetnames)
    if missing:
        result.errors.append(f"Hojas faltantes en el Excel: {', '.join(missing)}")
        return result

    # METADATA
    try:
        result.guide_create, result.guide_update = _parse_metadata(wb)
    except Exception as e:
        result.errors.append(f"Error en hoja METADATA: {e}")
        return result

    # RECOMENDADOS
    try:
        result.items += _parse_recomendados(wb)
    except Exception as e:
        result.warnings.append(f"Error en hoja RECOMENDADOS (ignorada): {e}")

    # INFLUENCERS
    try:
        result.items += _parse_influencers(wb)
    except Exception as e:
        result.warnings.append(f"Error en hoja INFLUENCERS (ignorada): {e}")

    # PERSONA_DEL_ANO
    try:
        persona_items, persona_extra = _parse_persona(wb)
        result.items += persona_items
        result.persona_extra = persona_extra
    except Exception as e:
        result.warnings.append(f"Error en hoja PERSONA_DEL_ANO (ignorada): {e}")

    if not result.items:
        result.warnings.append("No se encontraron items en el Excel.")

    return result
