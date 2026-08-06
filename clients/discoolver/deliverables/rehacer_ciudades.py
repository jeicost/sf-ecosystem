"""Rehace la hoja 'Estado por ciudad' del plan con los datos reales de la BBDD
de producción (CSV que pasó el CEO el 2026-08-06, exportado por Diego).

Criterio acordado con Diego:
  - STATE es el estado general; STATE=4 = recomendado (lo que se puede publicar).
  - ID_RECOMMENDED_BUSINESS_STATE es el estado de la traducción; se ignora, manda el primer STATE.
  - tab_url / main_url / gallery_urls = fotos de TAB, PRINCIPAL y GALERÍA.
Una tarjeta "lista" = viva + STATE=4 + foto principal + categoría.
"""
import csv, io, collections
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

CSV = '/Users/carlosjacoste/Developer/Claude/clients/discoolver/_snapshot/bbdd/produccion_recomendaciones_2026-08-06.csv'
XLSX = '/Users/carlosjacoste/Developer/Claude/clients/discoolver/deliverables/PLAN_LANZAMIENTO_2026-09-01.xlsx'

def F(s):
    try: return (s or '').encode('latin-1').decode('utf-8')
    except Exception: return s or ''

rows = [{k: F(v) for k, v in r.items()}
        for r in csv.DictReader(io.open(CSV, encoding='utf-8', errors='replace'), delimiter=';')]
vivos = [r for r in rows if r['deleted'] == '0']

stats = {}
for c in {r['city'] for r in vivos if r['city']}:
    rs = [r for r in vivos if r['city'] == c]
    r4 = [r for r in rs if r['STATE'] == '4']
    stats[c] = dict(
        vivas=len(rs),
        rec=len(r4),
        princ=sum(1 for r in r4 if r['main_url'].strip()),
        listas=sum(1 for r in r4 if r['main_url'].strip() and r['categoryName'].strip()),
        cola=sum(1 for r in rs if r['STATE'] == '3'),
        borr=sum(1 for r in rs if r['STATE'] == '1'),
        sinfoto=sum(1 for r in r4 if not (r['main_url'].strip() or r['tab_url'].strip() or r['gallery_urls'].strip())),
    )

# Qué falta / riesgo, escrito a mano por ciudad — el dato manda, el juicio lo pongo yo
NOTAS = {
 'Madrid': ('ola 1', 'bajo',
   'Nada de contenido: 858 tarjetas listas, 14× el mínimo de una guía. El trabajo es EDITORIAL, no de '
   'recolección: elegir 60-80, ordenarlas por barrio y vibra, y reescribir el copy con voz de guía. '
   'Sigue faltando fotografía con derechos de impresión si hay papel, el contrato con Cenando con Pablo '
   'y el checkout. 20 de las 975 recomendadas son fichas de IA: revisarlas a mano antes de que entren.'),
 'Barcelona': ('ola 1', 'bajo',
   '182 tarjetas listas, todas curadas por personas (cero IA). Da para guía sin recolectar nada más. '
   'Falta creador que la firme — hoy no hay ninguno asignado a Barcelona.'),
 'Ronda': ('ola 1', 'medio',
   '165 listas y es el destino donde discoolver ya está desplegado (tótems, POS). Guía + caso 360 en la '
   'misma plaza. Ojo: es un destino pequeño, el mercado de una guía de pago es limitado; encaja mejor '
   'como pieza de venta B2B que como producto B2C.'),
 'Málaga': ('ola 2', 'medio',
   '107 listas: suficiente para una guía ajustada. 85 fichas en STATE=1 sin revisar que podrían subir el '
   'número. Sin creador asignado.'),
 'Aranjuez': ('ola 2', 'alto',
   'Justo en el mínimo (64). Y es una escapada de día desde Madrid, no una ciudad de guía propia: encaja '
   'mejor como capítulo dentro de Madrid que como edición independiente.'),
 'Ibiza': ('ola 2', 'alto',
   '50 listas, por debajo del mínimo de 60. Faltan 10-30 fichas y 28 en cola de revisión que las darían. '
   'Producto muy estacional: lanzar en septiembre es lanzar al final de temporada.'),
 'Punta Cana': ('fuera de plan', 'medio',
   '128 listas y listas de verdad. No estaba en el plan del 1-sept y no hay creador ni mercado definido. '
   'Decisión de negocio: ¿el Caribe es mercado objetivo o es herencia del catálogo antiguo?'),
 'Santo Domingo': ('fuera de plan', 'medio',
   '75 listas. Mismo caso que Punta Cana: contenido listo sin plan comercial detrás.'),
 'Dubai': ('descartar', 'n/a',
   'Solo 11 fichas. No es una ciudad del catálogo, es un residuo.'),
 'Filipinas': ('limpiar', 'n/a',
   'NO ES UNA CIUDAD. Es un cajón de sastre con hoteles y bares internacionales (Burj Al Arab de Dubái, '
   'Ashford Castle en Irlanda, Marina Bay Sands de Singapur, bares de Londres y Roma). Las 50 fichas '
   'recomendadas hay que reasignarlas a su ciudad real o retirarlas: hoy ensucian cualquier recuento.'),
}

wb = load_workbook(XLSX)
del wb['Estado por ciudad']
ws = wb.create_sheet('Estado por ciudad', wb.sheetnames.index('Promesas web') + 1)

TITULO = Font(name='Arial', size=11, bold=True, color='FFFFFF')
HDRFILL = PatternFill('solid', start_color='1F2430')
NORMAL = Font(name='Arial', size=10)
BOLD = Font(name='Arial', size=10, bold=True)
WRAP = Alignment(wrap_text=True, vertical='top')
CENTER = Alignment(horizontal='center', vertical='center')
THIN = Border(bottom=Side(style='thin', color='D9D9D9'))
OLA = {'ola 1': 'C6EFCE', 'ola 2': 'FFEB9C', 'fuera de plan': 'DDEBF7',
       'limpiar': 'FFC7CE', 'descartar': 'E7E6E6'}
RIESGO = {'bajo': 'C6EFCE', 'medio': 'FFEB9C', 'alto': 'FFC7CE'}

ws['A1'] = ('Fuente: export de la BBDD de producción (Diego, 6-ago-2026) — '
            '_snapshot/bbdd/produccion_recomendaciones_2026-08-06.csv · 7.253 fichas, 6.861 vivas. '
            'Criterio: STATE=4 = recomendado (el primer STATE manda, la traducción se ignora). '
            '"Tarjeta lista" = viva + STATE=4 + foto principal + categoría.')
ws['A1'].font = Font(name='Arial', size=9, italic=True, color='595959')
ws['A1'].alignment = WRAP
ws.merge_cells('A1:I1')
ws.row_dimensions[1].height = 42

cols = ['Ciudad', 'Fichas vivas', 'Recomendadas\n(STATE=4)', 'Con foto\nprincipal',
        'TARJETAS\nLISTAS', 'Cola de revisión\n(STATE=3)', 'Decisión', 'Riesgo',
        'Qué significa y qué falta']
for i, c in enumerate(cols, 1):
    cell = ws.cell(row=2, column=i, value=c)
    cell.font = TITULO; cell.fill = HDRFILL
    cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
ws.row_dimensions[2].height = 34

orden = ['Madrid', 'Barcelona', 'Ronda', 'Málaga', 'Aranjuez', 'Ibiza',
         'Punta Cana', 'Santo Domingo', 'Dubai', 'Filipinas']
r = 3
for c in orden:
    s = stats[c]; decision, riesgo, nota = NOTAS[c]
    vals = [c, s['vivas'], s['rec'], s['princ'], s['listas'], s['cola'], decision, riesgo, nota]
    for i, v in enumerate(vals, 1):
        cell = ws.cell(row=r, column=i, value=v)
        cell.font = BOLD if i in (1, 5) else NORMAL
        cell.border = THIN
        cell.alignment = WRAP if i == 9 else (CENTER if i > 1 else Alignment(vertical='top'))
    ws.cell(row=r, column=7).fill = PatternFill('solid', start_color=OLA[decision])
    if riesgo in RIESGO:
        ws.cell(row=r, column=8).fill = PatternFill('solid', start_color=RIESGO[riesgo])
    if s['listas'] >= 60:
        ws.cell(row=r, column=5).fill = PatternFill('solid', start_color='C6EFCE')
    else:
        ws.cell(row=r, column=5).fill = PatternFill('solid', start_color='FFC7CE')
    ws.row_dimensions[r].height = 118
    r += 1

r += 1
ws.cell(row=r, column=1, value='LO QUE NO ESTÁ EN LA BBDD').font = Font(name='Arial', size=11, bold=True, color='C00000')
r += 1
for txt in [
  'Bangkok: CERO fichas. El plan del 1-sept la daba como una de las ocho plazas. Se construye desde nada — '
  'no hay atajo y no cabe en 26 días.',
  'Sevilla, Valencia, Bilbao, Granada, San Sebastián, Córdoba: CERO fichas. Ninguna de las ciudades '
  'españolas "obvias" existe en el catálogo.',
  'Las "7 ciudades de España" del plan no existen: solo hay 5 españolas con contenido real (Madrid, '
  'Barcelona, Ronda, Málaga, Aranjuez) más Ibiza por debajo del mínimo. El objetivo hay que reescribirlo '
  'sobre lo que hay, no al revés.',
]:
    cell = ws.cell(row=r, column=1, value=txt)
    cell.font = NORMAL; cell.alignment = WRAP
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
    ws.row_dimensions[r].height = 30
    r += 1

r += 1
ws.cell(row=r, column=1, value='EL CUELLO DE BOTELLA REAL').font = Font(name='Arial', size=11, bold=True, color='C00000')
r += 1
ia_vivas = sum(1 for x in vivos if x['rawId'].startswith('ia_'))
cola3 = [x for x in vivos if x['STATE'] == '3']
ia3 = sum(1 for x in cola3 if x['rawId'].startswith('ia_'))
foto3 = sum(1 for x in cola3 if x['main_url'].strip() or x['tab_url'].strip() or x['gallery_urls'].strip())
for txt in [
  f'La máquina de recolección SÍ funciona: {ia_vivas:,} de las {len(vivos):,} fichas vivas '
  f'({ia_vivas/len(vivos)*100:.0f}%) las generó la IA. Corregido: lo que parecía "el curador nunca ha '
  'arrancado" era el dev.db local del editor, no producción.'.replace(',', '.'),
  f'Lo que NO funciona es la revisión: {len(cola3):,} fichas esperando en STATE=3, el {ia3/len(cola3)*100:.0f}% '
  f'generadas por IA y solo el {foto3/len(cola3)*100:.0f}% con alguna foto. Recolectar es gratis; revisar y '
  'fotografiar es el trabajo caro, y es el que no se ha hecho.'.replace(',', '.'),
  'Traducción al plan: el 1 de septiembre no lo decide cuántas fichas hay, lo deciden las horas de '
  'edición humana y la fotografía con derechos. Madrid, Barcelona y Ronda ya tienen material de sobra; '
  'lo que falta es editor, fotógrafo y creador que firme.',
  'Limpieza rápida antes de publicar nada: 56 fichas recomendadas sin ninguna foto, 15 sin categoría, '
  '2 con etiquetas HTML <b> en el nombre, 6 nombres duplicados y 4 fichas de prueba de Diego.',
]:
    cell = ws.cell(row=r, column=1, value=txt)
    cell.font = NORMAL; cell.alignment = WRAP
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
    ws.row_dimensions[r].height = 32
    r += 1

for col, w in zip('ABCDEFGHI', [15, 12, 13, 11, 11, 14, 14, 9, 78]):
    ws.column_dimensions[col].width = w
ws.freeze_panes = 'A3'
ws.auto_filter.ref = f'A2:I{2 + len(orden)}'
ws.sheet_view.showGridLines = False

wb.save(XLSX)
print('OK — hoja rehecha con', len(orden), 'ciudades reales')
